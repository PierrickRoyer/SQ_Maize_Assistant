import pandas as pd
from lxml import etree
from xml.dom import minidom
from collections import defaultdict
import xml.etree.ElementTree as ET
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import matplotlib.pyplot as plt
import seaborn as sns

from Constant import varJugurta


def read_xml_to_df(file_path):
    parser = etree.XMLParser(remove_blank_text=True)  # To preserve formatting
    tree = etree.parse(file_path, parser)
    root = tree.getroot()

    # Extract data into a list for DataFrame creation
    data = []
    for item in root.xpath('//CropParameterItem'):
        param_name = item.get('name')
        for param in item.xpath('ParamValue/Item'):
            key = param.xpath('Key/string')[0].text
            value = param.xpath('Value/double')[0].text
            data.append([param_name, key, float(value)])

    # Create DataFrame
    df = pd.DataFrame(data, columns=['Variety', 'Parameter', 'Value'])

    return df, tree
# Function to rewrite the XML while preserving the namespaces and XML declaration
def rewrite_xml(tree, output_file):
    # Open the file in write mode
    with open(output_file, 'wb') as f:
        # Manually write the XML declaration without encoding
        f.write(b'<?xml version="1.0"?>\n')

        # Write the rest of the XML tree, without specifying encoding
        tree.write(f, pretty_print=True, xml_declaration=False)

def read_genotype_parameters(file_path):
    parser = etree.XMLParser(remove_blank_text=True)
    tree = etree.parse(file_path, parser)
    root = tree.getroot()

    # Dictionary to store parameters for each genotype
    genotype_params = defaultdict(dict)

    # Loop through each genotype (CropParameterItem)
    for item in root.xpath('//CropParameterItem'):
        genotype = item.get('name')  # Get the name of the genotype
        # Loop through each parameter under ParamValue
        for param in item.xpath('ParamValue/Item'):
            key = param.xpath('Key/string')[0].text  # Parameter name
            value = float(param.xpath('Value/double')[0].text)  # Parameter value as float
            genotype_params[genotype][key] = value

    return genotype_params,tree
# Function to find common parameters across genotypes with the same value
def find_common_parameters(genotype_params):
    # Dictionary to track potential common parameters and their value
    common_params = {}

    # Get the list of all genotypes
    genotypes = list(genotype_params.keys())

    # Use the first genotype as a reference
    reference_genotype = genotypes[0]

    # Loop through all parameters of the reference genotype
    for param_key, param_value in genotype_params[reference_genotype].items():
        # Check if all other genotypes have the same value for this parameter
        if all(genotype_params[genotype].get(param_key) == param_value for genotype in genotypes):
            common_params[param_key] = param_value

    return common_params
def process_all_sqvarm_files(root_dir):
    result_dict = {}

    # Walk through the root directory and its subdirectories
    for dirpath, _, filenames in os.walk(root_dir):
        for filename in filenames:
            if filename.endswith('.sqvarm'):
                file_path = os.path.join(dirpath, filename)
                print(f"Processing: {file_path}")

                # Read the parameters for each genotype
                genotype_params, _ = read_genotype_parameters(file_path)

                # Find common parameters with the same value across all genotypes
                common_params = find_common_parameters(genotype_params)

                # Store the results in the result_dict using the filename (without extension) as the key
                file_key = os.path.splitext(os.path.basename(file_path))[0]
                result_dict[file_key] = common_params

    return result_dict
# Function to read the XML file, remove a specific parameter, and return the modified tree
def read_and_remove_parameter(file_path, param_to_remove):
    parser = etree.XMLParser(remove_blank_text=True)
    tree = etree.parse(file_path, parser)
    root = tree.getroot()

    # Loop through each genotype (CropParameterItem)
    for item in root.xpath('//CropParameterItem'):
        # Loop through each parameter under ParamValue and remove the one to delete
        for param in item.xpath('ParamValue/Item'):
            key = param.xpath('Key/string')[0].text  # Parameter name
            if key == param_to_remove:
                parent = param.getparent()
                parent.remove(param)
                break  # Assuming you only want to remove one instance

    return tree
# Function to convert the result dictionary to a DataFrame
def convert_result_to_df(result_dict):
    # Create a list of Series to concatenate later
    param_series_list = []

    # Iterate through the result dictionary and add each file's common parameters
    for file_name, common_params in result_dict.items():
        # Create a Series with the file name as the index
        param_series = pd.Series(common_params, name=file_name)
        param_series_list.append(param_series)

    # Concatenate all Series into a DataFrame
    df = pd.concat(param_series_list, axis=1).T

    # Reset index to make 'File' a column
    df.reset_index(inplace=True)
    df.rename(columns={'index': 'File'}, inplace=True)

    return df
# Function to subselect columns where all values are the same and there are no NA
def subselect_columns(df):
    # Filter columns where all values are the same and not NA
    same_value_columns = df.nunique().eq(1)  # Check for uniqueness
    no_na_columns = df.notna().all()         # Check for NA

    # Combine conditions to get the desired columns
    selected_columns = same_value_columns & no_na_columns

    # Subselect the DataFrame
    result_df = df.loc[:, selected_columns]

    return result_df

def create_excluded_columns_df(original_df, filtered_df):
    # Get the columns to exclude
    excluded_columns = filtered_df.columns

    # Create a new DataFrame with columns not in filtered_df
    excluded_df = original_df.drop(columns=excluded_columns, errors='ignore')

    return excluded_df
# Function to remove parameters from B that are not in A and track the removed params
def remove_extra_parameters_in_B(tree_B, params_A):
    root_B = tree_B.getroot()
    removed_params = set()

    # Iterate through each genotype in file B
    for item in root_B.xpath('//CropParameterItem'):
        for param in item.xpath('ParamValue/Item'):
            key = param.xpath('Key/string')[0].text
            if key not in params_A[next(iter(params_A))]:  # If param not in A, remove it
                parent = param.getparent()
                parent.remove(param)
                removed_params.add(key)  # Track removed parameter

    return tree_B, removed_params
# Function to add missing parameters from A to B with the average value from A
def add_missing_parameters_in_B(tree_B, avg_values_to_add):
    root_B = tree_B.getroot()
    added_params = {}

    # Iterate through each genotype in file B and add missing parameters
    for item in root_B.xpath('//CropParameterItem'):
        for param_name, avg_value in avg_values_to_add.items():
            # Check if the parameter is missing in the genotype
            existing_keys = {param.xpath('Key/string')[0].text for param in item.xpath('ParamValue/Item')}
            if param_name not in existing_keys:
                # Create a new parameter element and add it to the genotype
                new_param = etree.Element("Item")
                key = etree.SubElement(new_param, "Key")
                key_string = etree.SubElement(key, "string")
                key_string.text = param_name

                value = etree.SubElement(new_param, "Value")
                value_double = etree.SubElement(value, "double")
                value_double.text = str(avg_value)

                # Append the new parameter to the ParamValue section
                item.find('ParamValue').append(new_param)
                added_params[param_name] = avg_value  # Track added parameter

    return tree_B, added_params

def log_modifications(file_path, removed_params, added_params):
    with open(file_path, 'w') as log_file:
        if added_params:
            log_file.write("Parameters added:\n")
            for param, value in added_params.items():
                log_file.write(f"{param}, {value}\n")

        if removed_params:
            log_file.write("\nParameters removed:\n")
            for param in removed_params:
                log_file.write(f"{param}\n")
# New function to treat NFinal param
def treat_NFinal(tree_B):
    root_B = tree_B.getroot()

    # Iterate through each genotype in file B
    for item in root_B.xpath('//CropParameterItem'):
        for param in item.xpath('ParamValue/Item'):
            key = param.xpath('Key/string')[0].text
            if key == 'Nfinal':  # Apply treatment only to NFinal parameter
                nfinal_value = float(param.xpath('Value/double')[0].text)

                # Check if the decimal part of NFinal is smaller than 0.5
                if nfinal_value - int(nfinal_value) < 0.5:
                    nfinal_value -= 0.5  # Subtract 0.5 if the condition is met
                    # Update the value in the XML
                    param.xpath('Value/double')[0].text = str(nfinal_value)

    return tree_B

def collect_sqvarm_data(root_dir):
    data = []  # To store information about each database, version, and file
    for database_version in os.listdir(root_dir):
        db_path = os.path.join(root_dir, database_version)
        if os.path.isdir(db_path):
            # Split the database and version from directory name
            if '_' in database_version:
                database, version = database_version.split('_', 1)
            else:
                database = database_version
                version = 'Unknown'  # Fallback if version is not provided

            # Check for .sqvarm files directly in the database directory (version root)
            root_sqvarms = [f for f in os.listdir(db_path) if f.endswith('.sqvarm')]
            if root_sqvarms:
                for file in root_sqvarms:
                    data.append({
                        'Database': database,
                        'Version': version,
                        'File': file,
                        'FilePath': os.path.join(db_path, file)
                    })

            # Now check for subdirectories representing model versions
            for subdir_version in os.listdir(db_path):
                version_path = os.path.join(db_path, subdir_version)
                if os.path.isdir(version_path):
                    # Collect .sqvarm files in subdirectories (versions)
                    for file in os.listdir(version_path):
                        if file.endswith('.sqvarm'):
                            data.append({
                                'Database': database,
                                'Version': subdir_version,
                                'File': file,
                                'FilePath': os.path.join(version_path, file)
                            })
    return pd.DataFrame(data)

def compare_sqvarm_files(sqvarm_df):
    # You can use your comparison functions here to check for differences
    comparison_results = []

    for db in sqvarm_df['Database'].unique():
        db_versions = sqvarm_df[sqvarm_df['Database'] == db]

        # Compare versions within the same database
        for version in db_versions['Version'].unique():
            version_files = db_versions[db_versions['Version'] == version]

            # Apply your parameter comparison logic here:
            # Compare the parameters of each sqvarm file in this version
            # You can add the logic to compare the sets as per your previous code
            comparison_results.append({
                'Database': db,
                'Version': version,
                'File': ','.join(version_files['File'].tolist()),
                'Differences': 'Differences placeholder'  # Insert real differences here
            })

    return pd.DataFrame(comparison_results)

def sync_files_A_and_B(file_A, file_B, output_file_B, log_file_path,targetVersion):
    # Read parameters from both files
    params_A, _ = read_genotype_parameters(file_A)
    params_B, tree_B = read_genotype_parameters(file_B)

    # Remove parameters from B that are not in A and track changes
    modified_tree_B, removed_params = remove_extra_parameters_in_B(tree_B, params_A)

    # Calculate the average values of parameters in A that need to be added to B
    avg_values_to_add = {}
    for param in (set(params_A[next(iter(params_A))].keys()) - set(params_B[next(iter(params_B))].keys())):
        avg_values_to_add[param] = sum(genotype[param] for genotype in params_A.values() if param in genotype) / len(params_A)

    # Add missing parameters to B from A and track changes
    modified_tree_B, added_params = add_missing_parameters_in_B(modified_tree_B, avg_values_to_add)

    # Apply NFinal treatment to B
    if targetVersion =='2.0.0_Nfinal':
        modified_tree_B = treat_NFinal(modified_tree_B)

    os.makedirs(os.path.dirname(output_file_B), exist_ok=True)
    #os.makedirs(os.path.dirname(log_file_path), exist_ok = True)
    # Rewrite the modified XML tree for B to the specified output file
    rewrite_xml(modified_tree_B, output_file_B)

    # Log modifications
    #log_modifications(log_file_path, removed_params, added_params)



def generate_param(df, output_file, switch_var_nonvar):
    # Write XML header based on the variety or non-variety parameter
    if switch_var_nonvar:
        root = ET.Element('MaizeVarietyFile', {
            'xmlns:xsd': "http://www.w3.org/2001/XMLSchema",
            'xmlns:xsi': "http://www.w3.org/2001/XMLSchema-instance"
        })
        items_array = ET.SubElement(root, 'ItemsArray')
    else:
        root = ET.Element('MaizeNonVarietyFile', {
            'xmlns:xsd': "http://www.w3.org/2001/XMLSchema",
            'xmlns:xsi': "http://www.w3.org/2001/XMLSchema-instance"
        })
        items_array = ET.SubElement(root, 'ItemsArray')

    # Loop through each column (each genotype)
    for col in df.columns[1:]:
        crop_item = ET.SubElement(items_array, 'CropParameterItem', {'name': col})

        # Create a single ParamValue for the current genotype
        param_value_elem = ET.SubElement(crop_item, 'ParamValue')

        # Loop through each parameter for the current genotype
        for i, param in enumerate(df['Parameter']):
            param_value = df[col][i]

            # Create Item for each parameter
            item_elem = ET.SubElement(param_value_elem, 'Item')

            key_elem = ET.SubElement(item_elem, 'Key')
            key_sub_elem = ET.SubElement(key_elem, 'string')
            key_sub_elem.text = param

            value_elem = ET.SubElement(item_elem, 'Value')
            value_double_elem = ET.SubElement(value_elem, 'double')
            value_double_elem.text = str(param_value)

    # Pretty print the XML
    xml_str = ET.tostring(root, encoding='utf-8', method='xml')
    parsed_xml = minidom.parseString(xml_str)
    pretty_xml_str = parsed_xml.toprettyxml(indent="  ")

    # Write the pretty XML string to the output file
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(pretty_xml_str)

    return pretty_xml_str


def sync_non_variety_files(file_A, file_B, output_file_B, log_file_path):
    # Read parameters from both files
    params_A, tree_B = read_non_variety_parameters(file_A)
    params_B, _ = read_non_variety_parameters(file_B)

    # Dictionary to track average values to add for each genotype
    avg_values_to_add = {}

    # Iterate through each CropParameterItem in params_A
    for genotype_name, parameters in params_A.items():
        if genotype_name not in params_B:
            continue  # Skip if the genotype is not in B

        # Calculate missing parameters
        for param_name, value in parameters.items():
            if param_name not in params_B[genotype_name]:
                # Initialize if not yet added
                if param_name not in avg_values_to_add:
                    avg_values_to_add[param_name] = 0
                avg_values_to_add[param_name] += value

    # Calculate averages if applicable
    for param_name in avg_values_to_add:
        avg_values_to_add[param_name] /= len(params_A)  # Average across all genotypes in A

    # Add missing parameters to each CropParameterItem in B from A
    modified_tree_B, added_params = add_missing_parameters_in_non_variety(tree_B, params_A)

    # Ensure directories exist for output and log files
    os.makedirs(os.path.dirname(output_file_B), exist_ok=True)
    os.makedirs(os.path.dirname(log_file_path), exist_ok=True)

    # Rewrite the modified XML tree for B to the specified output file
    modified_tree_B.write(output_file_B, pretty_print=True, xml_declaration=True, encoding='UTF-8')

    # Log modifications
    log_modifications(log_file_path, [], added_params)




def read_non_variety_parameters(file_path):
    """Reads the non-variety sqparm file and returns parameters along with the parsed XML tree."""
    tree = etree.parse(file_path)
    root = tree.getroot()

    params = {}
    # Extract parameters from each CropParameterItem
    for item in root.findall('.//CropParameterItem'):
        genotype_name = item.get('name')
        params[genotype_name] = {}

        for param_item in item.findall('.//Item'):
            key = param_item.find('.//Key/string').text
            value_element = param_item.find('.//Value/double')
            if value_element is not None:
                value = value_element.text
                if value is not None:
                    try:
                        params[genotype_name][key] = float(value)
                    except ValueError:
                        params[genotype_name][key] = value  # Keep non-numeric values as they are
    return params, tree



def add_missing_parameters_in_non_variety(tree, params_A):
    """ Add missing parameters from A to the XML tree. """
    added_params = []
    root = tree.getroot()

    # Iterate through CropParameterItems in the XML tree
    for item in root.findall('.//CropParameterItem'):
        genotype_name = item.get('name')

        # If genotype exists in A, check for missing parameters
        if genotype_name in params_A:
            for param_name, param_value in params_A[genotype_name].items():
                if item.find(f".//ParamValue[@Key='{param_name}']") is None:
                    # Create a new ParamValue element
                    new_param = etree.Element('ParamValue', Key=param_name, Value=str(param_value))
                    item.append(new_param)  # Add to CropParameterItem
                    added_params.append((genotype_name, param_name, param_value))

    return tree, added_params

def log_modifications(log_file_path, removed_params, added_params):
    """ Log modifications made during syncing. """
    with open(log_file_path, 'a') as log_file:
        log_file.write("Added Parameters:\n")
        for genotype, param, value in added_params:
            log_file.write(f"Genotype: {genotype}, Added Param: {param}, Value: {value}\n")

        log_file.write("Removed Parameters:\n")
        for genotype, param in removed_params:
            log_file.write(f"Genotype: {genotype}, Removed Param: {param}\n")
        log_file.write("\n")


def read_non_variety_parameters(file_path):
    """Reads the non-variety sqparm file and returns parameters along with the parsed XML tree."""
    tree = etree.parse(file_path)
    root = tree.getroot()

    params = {}
    # Extract parameters from each CropParameterItem
    for item in root.findall('.//CropParameterItem'):
        genotype_name = item.get('name')
        params[genotype_name] = {}

        for param_item in item.findall('.//Item'):
            key = param_item.find('.//Key/string').text
            value_element = param_item.find('.//Value/double')
            if value_element is not None:
                value = value_element.text
                if value is not None:
                    try:
                        params[genotype_name][key] = float(value)
                    except ValueError:
                        params[genotype_name][key] = value  # Keep non-numeric values as they are
    return params, tree


def edit_parameters_in_tree(tree, df):
    """Edits the parameters in the XML tree based on the DataFrame."""
    root = tree.getroot()

    # Extract the genotype name (column name) from the DataFrame
    genotype = df.columns[1]  # Assuming the first column is 'Parameter', and the second column is the genotype name

    # Find the correct CropParameterItem for the genotype
    item = root.find(f".//CropParameterItem[@name='{genotype}']")
    if item is not None:
        # Iterate over the DataFrame rows and update corresponding XML values
        for _, row in df.iterrows():
            param_key = row['Parameter']
            new_value = row[genotype]

            # Find the corresponding Item with the matching Key string
            # Corrected XPath query to find the Key string
            param_value_item = item.xpath(f".//Item[Key/string[text()='{param_key}']]")

            if param_value_item:
                value_element = param_value_item[0].find('.//Value/double')
                if value_element is not None:
                    value_element.text = str(new_value)

def save_tree_to_file(tree, output_file_path):
    """Save the modified XML tree to a new file."""
    tree.write(output_file_path, pretty_print=True, xml_declaration=True, encoding='UTF-8')


def sync_non_variety_file_with_df(xml_file_path, df, output_file_path):
    """Reads the XML file, updates it based on the DataFrame, and writes the changes to a new file."""
    # Read the non-variety parameters and the XML tree
    params, tree = read_non_variety_parameters(xml_file_path)

    # Edit the tree based on the DataFrame
    edit_parameters_in_tree(tree, df)

    # Save the modified XML tree to a new file
    save_tree_to_file(tree, output_file_path)


def extract_parameters_from_sqvarm(sqvarm_file):
    """
    Extracts parameters from a .sqvarm file into a dictionary.
    """
    tree = ET.parse(sqvarm_file)
    root = tree.getroot()

    params = {}
    # Iterate through each CropParameterItem in the file
    for item in root.findall(".//CropParameterItem"):
        for param in item.find("ParamValue").findall("Item"):
            key = param.find("./Key/string").text
            value = float(param.find("./Value/double").text)
            params[key] = value

    return params

def summarize_parameters(params_list):
    """
    Summarize (min, max, avg) for a list of parameter dictionaries.
    """
    summary = {}
    # Loop through all parameters in the first parameter dictionary
    for param in params_list[0].keys():
        values = []
        for params in params_list:  # Get all values for this parameter across all dictionaries
            value = params.get(param, float('nan'))  # Get the value, default to NaN if not found
            if not pd.isna(value):  # Only consider valid (non-NaN) values
                values.append(value)

        if values:  # Only summarize if there are valid values
            summary[param] = {
                'min': min(values),
                'max': max(values),
                'avg': sum(values) / len(values)
            }

    return summary

def align_parameters(sqvarm_params, sqvarm_ref, sqparm_ref):
    """
    Align the parameters to the reference files and separate them into three categories:
    1. Parameters found in sqvarm_ref.
    2. Parameters found in sqparm_ref.
    3. Parameters found in neither.
    """
    sqvarm_group = {}
    sqparm_group = {}
    new_group = {}

    for param, value in sqvarm_params.items():
        if param in sqvarm_ref:
            sqvarm_group[param] = value
        elif param in sqparm_ref:
            sqparm_group[param] = value
        else:
            new_group[param] = value

    return sqvarm_group, sqparm_group, new_group

def summarize_sqvarm_files(sqvarm_dir, sqvarm_ref_file, sqparm_ref_file):
    """
    Summarize the min/max/avg of parameters across multiple .sqvarm files.
    """
    # Load the reference files
    sqvarm_ref = extract_parameters_from_sqvarm(sqvarm_ref_file)
    sqparm_ref = extract_parameters_from_sqvarm(sqparm_ref_file)

    # Iterate over the sqvarm files in the directory
    summaries = {
        'sqvarm_ref': {},
        'sqparm_ref': {},
        'new_params': {}
    }

    files_list = []

    for sqvarm_file in os.listdir(sqvarm_dir):
        if sqvarm_file.endswith(".sqvarm"):
            sqvarm_file_path = os.path.join(sqvarm_dir, sqvarm_file)
            params = extract_parameters_from_sqvarm(sqvarm_file_path)

            # Align parameters to references
            sqvarm_group, sqparm_group, new_group = align_parameters(params, sqvarm_ref, sqparm_ref)

            # Summarize each group and append to the lists
            summaries['sqvarm_ref'][sqvarm_file] = summarize_parameters([sqvarm_group])
            summaries['sqparm_ref'][sqvarm_file] = summarize_parameters([sqparm_group])
            summaries['new_params'][sqvarm_file] = summarize_parameters([new_group])

            files_list.append(sqvarm_file)

    # Convert summaries to DataFrames for easy representation
    df_sqvarm_ref = pd.DataFrame.from_dict(summaries['sqvarm_ref'], orient='index')
    df_sqparm_ref = pd.DataFrame.from_dict(summaries['sqparm_ref'], orient='index')
    df_new_params = pd.DataFrame.from_dict(summaries['new_params'], orient='index')

    return df_sqvarm_ref, df_sqparm_ref, df_new_params


def extract_parameters_to_df(sqvarm_file):
    """
    Extract parameters from a .sqvarm file and return as a DataFrame.
    """
    tree = ET.parse(sqvarm_file)
    root = tree.getroot()

    data = {}

    # Iterate through each CropParameterItem in the file
    for item in root.findall(".//CropParameterItem"):
        variety_name = item.get("name")  # Get the name of the crop variety
        params = item.find("ParamValue")

        # Iterate through each parameter item within ParamValue
        for param in params.findall("Item"):
            key = param.find("./Key/string").text
            value = float(param.find("./Value/double").text)

            # Store values in the data dictionary
            if variety_name not in data:
                data[variety_name] = {}
            data[variety_name][key] = value

    # Create DataFrame from the collected data
    df = pd.DataFrame.from_dict(data, orient='index')
    return df


def consolidate_dfs_and_calculate_avg(dfs):
    summary_data = {}

    # Iterate through each DataFrame in the list
    for df_name, df in dfs.items():
        for param in df.columns:
            # Initialize the dictionary for the parameter if it doesn't exist
            if param not in summary_data:
                summary_data[param] = {}

            # Calculate average value for the parameter in the current DataFrame
            avg_value = df[param].mean()
            # Check for constant values in the DataFrame
            is_constant = df[param].nunique() == 1

            # Store average value and constant status
            summary_data[param][f"{df_name}_avg"] = avg_value
            summary_data[param][f"{df_name}_isConstant"] = is_constant

    # Create the summary DataFrame
    summary_df = pd.DataFrame.from_dict(summary_data, orient='index')

    return summary_df


def apply_color_to_excel(writer, summary_df):
    """
    Apply colors to the Excel sheet based on the average values.

    Args:
        writer (pd.ExcelWriter): Pandas Excel writer.
        summary_df (pd.DataFrame): Summary DataFrame.
    """
    workbook = writer.book
    worksheet = writer.sheets['Summary']

    # Create fill objects for grey and color gradients
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Iterate over each parameter to find min and max values across the row
    for row_idx, param in enumerate(summary_df.index, start=2):  # start=2 to account for header row
        row_values = summary_df.loc[param]
        # Get the average values only
        avg_values = row_values[row_values.index.str.endswith('_avg')]

        # Get min and max values for this row
        min_val = avg_values.min()
        max_val = avg_values.max()

        for col_idx, avg_col in enumerate(avg_values.index, start=2):  # start=2 to account for header column
            avg_value = avg_values[avg_col]
            # Skip coloring if the value is NaN
            if pd.isna(avg_value):
                continue

            # Apply grey fill if the value is equal to min or max
            if avg_value == min_val or avg_value == max_val:
                worksheet.cell(row=row_idx, column=col_idx).fill = grey_fill
            else:
                # Scale color based on the value relative to the row's min and max
                if min_val != max_val:  # Avoid division by zero
                    scaled_val = (avg_value - min_val) / (max_val - min_val)
                else:
                    scaled_val = 0  # If all values are the same

                # Calculate color
                red = int(255 * (1 - scaled_val))  # Scale from 255 (red) to 0
                green = int(255 * scaled_val)  # Scale from 0 to 255 (green)
                fill_color = PatternFill(start_color=f"{red:02X}{green:02X}00", end_color=f"{red:02X}{green:02X}00",
                                         fill_type="solid")
                worksheet.cell(row=row_idx, column=col_idx).fill = fill_color


# Modify the sqvarm file by adding the parameters from sqparm
def add_sqparm_to_sqvarm(sqvarm_file, sqparm_file, output_file):
    # Read parameters from both files
    genotype_params, var_tree = read_genotype_parameters(sqvarm_file)
    non_variety_params, _ = read_non_variety_parameters(sqparm_file)
    print('"here')
    # Find all CropParameterItem elements in the sqvarm file
    for item in var_tree.xpath('//CropParameterItem'):
        genotype = item.get('name')  # The genotype name
        print(genotype)
        print(non_variety_params)
        # If genotype exists in the non-variety parameters, add them

        for key, value in non_variety_params['MAZ'].items():
            # Create a new parameter entry in the sqvarm XML structure
            param_value_elem = etree.SubElement(item.xpath('ParamValue')[0], 'Item')

            key_elem = etree.SubElement(param_value_elem, 'Key')
            key_string = etree.SubElement(key_elem, 'string')
            key_string.text = key

            value_elem = etree.SubElement(param_value_elem, 'Value')
            value_double = etree.SubElement(value_elem, 'double')
            value_double.text = str(value)

    # Rewrite the modified sqvarm tree to a new file
    rewrite_xml(var_tree, output_file)

# Example usage